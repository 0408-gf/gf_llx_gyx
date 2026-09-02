from __future__ import annotations

import csv
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

from models.match import Match

FIELD_ALIASES = {
    "match_time": ("match_time", "比赛时间", "日期时间", "时间"),
    "league": ("league", "联赛"),
    "home_team": ("home_team", "主队"),
    "away_team": ("away_team", "客队"),
    "home_odds": ("home_odds", "主胜赔率"),
    "draw_odds": ("draw_odds", "平局赔率"),
    "away_odds": ("away_odds", "客胜赔率"),
    "asian_line": ("asian_line", "亚洲让球盘口"),
    "asian_home_odds": ("asian_home_odds", "主队亚盘赔率"),
    "asian_away_odds": ("asian_away_odds", "客队亚盘赔率"),
    "total_line": ("total_line", "大小球盘口"),
    "over_odds": ("over_odds", "大球赔率"),
    "under_odds": ("under_odds", "小球赔率"),
    "source": ("source", "数据来源", "来源"),
    "updated_at": ("updated_at", "更新时间"),
}
NUMBERS = {"home_odds", "draw_odds", "away_odds", "asian_line", "asian_home_odds", "asian_away_odds", "total_line", "over_odds", "under_odds"}
REQUIRED = ("match_time", "league", "home_team", "away_team", "source")
DISPLAY_NAMES = {"match_time": "时间", "league": "联赛", "home_team": "主队", "away_team": "客队", "source": "来源"}
NUMBER_DISPLAY_NAMES = {
    "home_odds": "主胜赔率", "draw_odds": "平局赔率", "away_odds": "客胜赔率",
    "asian_line": "亚洲让球盘口", "asian_home_odds": "主队亚盘赔率",
    "asian_away_odds": "客队亚盘赔率", "total_line": "大小球盘口",
    "over_odds": "大球赔率", "under_odds": "小球赔率",
}
EXAMPLE = "时间,联赛,主队,客队,来源\n2026-01-01 19:30,测试联赛,甲队,乙队,本地文件"
_ALIAS_TO_FIELD = {alias.strip().lstrip("\ufeff").casefold(): field for field, aliases in FIELD_ALIASES.items() for alias in aliases}


def _clean_header(value: Any) -> str:
    return "" if value is None else str(value).strip().lstrip("\ufeff").strip()


def _headers(values: Iterable[Any], context: str = "文件") -> list[str]:
    headers = [_clean_header(value) for value in values]
    if not headers or not any(headers):
        raise ValueError(f"{context}表头为空。必填字段：时间、联赛、主队、客队、来源。\n可复制示例：\n{EXAMPLE}")
    if any(not header for header in headers):
        raise ValueError(f"{context}表头包含空字段名，请删除空列或补全字段名。")
    normalized = [header.casefold() for header in headers]
    if len(normalized) != len(set(normalized)):
        raise ValueError(f"{context}表头包含重复字段，请确保每个字段只出现一次。")
    meanings = [_ALIAS_TO_FIELD.get(header) for header in normalized]
    recognized = [meaning for meaning in meanings if meaning]
    if len(recognized) != len(set(recognized)):
        raise ValueError(f"{context}表头包含含义重复的字段（例如“时间”和“比赛时间”），请只保留一个。")
    missing = [DISPLAY_NAMES[field] for field in REQUIRED if field not in recognized]
    if missing:
        raise ValueError(f"{context}表头缺少必填字段：{'、'.join(missing)}。\n可复制示例：\n{EXAMPLE}")
    return [meaning or header for meaning, header in zip(meanings, headers)]


def _date(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError(f"无法识别日期：{text}")


def _mapped(row: dict[str, Any], line: int) -> Match:
    data = {key: row.get(key) for key in FIELD_ALIASES}
    missing = [DISPLAY_NAMES[field] for field in REQUIRED if data[field] is None or not str(data[field]).strip()]
    if missing:
        raise ValueError(f"第 {line} 行缺少必填字段：{'、'.join(missing)}")
    for key in NUMBERS:
        if data[key] is not None and str(data[key]).strip():
            try:
                data[key] = float(data[key])
            except (TypeError, ValueError):
                raise ValueError(f"第 {line} 行的{NUMBER_DISPLAY_NAMES[key]}不是有效数字。") from None
        else:
            data[key] = None
    data["match_time"] = _date(data["match_time"])
    data["updated_at"] = _date(data["updated_at"]) if data["updated_at"] and str(data["updated_at"]).strip() else datetime.now()
    return Match(**data)


def _matches_from_matrix(matrix: list[list[Any]], context: str = "文件") -> list[Match]:
    matrix = [list(row) for row in matrix if any(value is not None and str(value).strip() for value in row)]
    if not matrix:
        raise ValueError(f"{context}为空，没有表头和比赛数据。")
    headers = _headers(matrix[0], context)
    matches: list[Match] = []
    for line, values in enumerate(matrix[1:], 2):
        if len(values) != len(headers):
            raise ValueError(f"第 {line} 行列数与表头不一致：应为 {len(headers)} 列，实际为 {len(values)} 列。")
        matches.append(_mapped(dict(zip(headers, values)), line))
    if not matches:
        raise ValueError(f"{context}只有表头，没有比赛数据。")
    return matches


def _read_text(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            pass
    raise ValueError("无法识别文件编码，请将文件保存为 UTF-8（推荐）或 GB18030 后重试。")


def _delimited(path: Path) -> list[Match]:
    text = _read_text(path)
    if not text.strip():
        raise ValueError("文件为空，没有表头和比赛数据。")
    first = next((line for line in text.splitlines() if line.strip()), "")
    candidates = ("\t", ",", ";", "|")
    delimiter = max(candidates, key=first.count)
    if first.count(delimiter) == 0:
        raise ValueError(f"TXT/CSV 必须是含表头的结构化表格，支持制表符、逗号、分号或竖线分隔；不能导入自然语言。\n必填字段：时间、联赛、主队、客队、来源。\n可复制示例：\n{EXAMPLE}")
    try:
        matrix = list(csv.reader(text.splitlines(), delimiter=delimiter, strict=True))
    except csv.Error as exc:
        raise ValueError(f"TXT/CSV 格式错误：{exc}") from None
    return _matches_from_matrix(matrix, "TXT/CSV 文件")


class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[tuple[list[list[str]], bool]] = []
        self._depth = 0
        self._rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None
        self._merged = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.casefold()
        if tag == "table":
            self._depth += 1
            if self._depth == 1:
                self._rows, self._merged = [], False
        elif self._depth == 1 and tag == "tr":
            self._row = []
        elif self._depth == 1 and tag in ("th", "td") and self._row is not None:
            attr = {key.casefold(): value for key, value in attrs}
            if "rowspan" in attr or "colspan" in attr:
                self._merged = True
            self._cell = []
        elif self._depth == 1 and tag == "br" and self._cell is not None:
            self._cell.append(" ")

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.casefold()
        if self._depth == 1 and tag in ("th", "td") and self._cell is not None and self._row is not None:
            self._row.append(" ".join("".join(self._cell).split()))
            self._cell = None
        elif self._depth == 1 and tag == "tr" and self._row is not None:
            if self._row:
                self._rows.append(self._row)
            self._row = None
        elif tag == "table" and self._depth:
            if self._depth == 1:
                self.tables.append((self._rows, self._merged))
            self._depth -= 1


def _html(path: Path) -> list[Match]:
    parser = _TableParser()
    parser.feed(_read_text(path))
    parser.close()
    if not parser.tables:
        raise ValueError("HTML 中没有真实的 <table> 表格。脚本动态生成的占位内容不会执行，也不会联网加载。")
    candidates: list[tuple[int, list[list[str]]]] = []
    saw_merged = False
    for matrix, merged in parser.tables:
        saw_merged |= merged
        if not matrix or merged:
            continue
        try:
            normalized = _headers(matrix[0], "HTML 表格")
        except ValueError:
            continue
        candidates.append((sum(header in FIELD_ALIASES for header in normalized), matrix))
    if not candidates:
        if saw_merged:
            raise ValueError("HTML 表格含 rowspan/colspan 合并单元格，无法可靠展开，为防止列错位已拒绝导入。请先取消合并单元格。")
        raise ValueError(f"HTML 中没有必填字段齐全的比赛表格。必填字段：时间、联赛、主队、客队、来源。\n可复制示例：\n{EXAMPLE}")
    return _matches_from_matrix(max(candidates, key=lambda item: item[0])[1], "HTML 表格")


def _xlsx(path: Path) -> list[Match]:
    from zipfile import BadZipFile, is_zipfile

    advice = "无法读取 XLSX：文件可能损坏、被占用，或只是把其他文件改成了 .xlsx 后缀。请恢复正确扩展名后按 HTML/TXT/CSV 导入，或在 Excel/WPS 中真正“另存为 Excel 工作簿 (*.xlsx)”；仅改后缀不等于转换。"
    try:
        if not is_zipfile(path):
            raise ValueError(advice)
    except (PermissionError, OSError) as exc:
        raise ValueError(advice) from exc

    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if not workbook.worksheets:
            raise ValueError("Excel 工作簿中没有工作表。")
        matrix = [list(row) for row in workbook.active.iter_rows(values_only=True)]
        if matrix and not any(value is not None and str(value).strip() for value in matrix[0]):
            raise ValueError("Excel 工作表首行为空；请将字段表头放在第一行。")
        return _matches_from_matrix(matrix, "Excel 工作表")
    except (BadZipFile, InvalidFileException, PermissionError, OSError) as exc:
        raise ValueError(advice) from exc


def import_matches(path: str | Path) -> list[Match]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".csv", ".txt"):
        return _delimited(path)
    if suffix == ".xlsx":
        return _xlsx(path)
    if suffix in (".html", ".htm"):
        return _html(path)
    if suffix in (".jpg", ".jpeg", ".png"):
        raise ValueError("截图图片必须通过界面的“截图识别预览/校对”导入，确认前不会写入数据库。")
    raise ValueError("仅支持 CSV、TXT、XLSX、HTML、HTM、JPG、JPEG 和 PNG 本地文件；不会执行脚本、联网或抓取网站。")
